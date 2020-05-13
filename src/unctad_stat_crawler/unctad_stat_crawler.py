import requests
import xml.etree.ElementTree as ET
import time
import json
from collections import namedtuple
from urllib.parse import quote, unquote
from os import mkdir
from os.path import isdir
from os.path import split as splitPath
from os.path import join as joinPath
from openpyxl import Workbook
from random import random
from .itemcode_getter import ItemCode_Getter

# Global PARAM ---------------------------------------------------------
## supported data
SUPPORT_DATA = ["import", "export", "both"]

## request info
GET_DATA_INFO_PATH = "./config/getData.json"

## namedtuple list
Request_info = namedtuple("Request_info", ["URL", "headers", "postData"])
Request_param = namedtuple("Request_param", ["category", 
                                             "sWD_ReportId", "sWD_TableId", 
                                             "sheet_name", "sheet_title"])
## request_info
REQUEST_PARAM = dict()
REQUEST_PARAM["export"] = Request_param(category="export",
                                        sWD_ReportId='182117', 
                                        sWD_TableId='1495', 
                                        sheet_name="科技產品出口",
                                        sheet_title="Merchandise trade matrix – " + \
                                                    "product groups, exports in thousands " + \
                                                    "of United States dollars, ")

REQUEST_PARAM["import"] = Request_param(category="import", 
                                        sWD_ReportId='180622', 
                                        sWD_TableId='1494', 
                                        sheet_name="科技產品進口", 
                                        sheet_title="Merchandise trade matrix – " + \
                                                    "product groups, imports in thousands " + \
                                                    "of United States dollars, ")

## supported data
CATEGORY = dict()
CATEGORY["import"] = tuple([REQUEST_PARAM["import"]])
CATEGORY["export"] = tuple([REQUEST_PARAM["export"]])
CATEGORY["both"] = (REQUEST_PARAM["export"], REQUEST_PARAM["import"])
# ----------------------------------------------------------------------

class UNCTAD_STAT_Crawler:
    def __init__(self, delay_max = 1, update_code = False):
        '''
        @param delay_max:
        To mimic human's action, determine the gap between
        every request to the target. The default value is 1,
        and our program will select a number on the interval
        [0, delay_max]
        @param update_code:
            Before we start making a request to get the specific tables from
            the target website, it is of inevitable that we obtain the itemCode
            of each item, such as 'economy', 'product', and 'partner',  in advance.  
            There's 2 ways for us to get such the information.
            1. from the local file, /src/config/<item>Code.json
            2. request from the https://unctadstat.unctad.org/
            The parameter 'update_code' is the thing that you can force the
            crawler to update the information and save it to the /src/config.
        '''
        self.delay_max = delay_max
                
        self.economyCode = ItemCode_Getter("economy")        
        self.productCode = ItemCode_Getter("product")
        self.partnerCode = ItemCode_Getter("partner")
        
        if(update_code):
            self.update_all_code()
        
        self.rq_info = None
        self._read_request_info()# assign value to self.rq_info
        
    def _read_request_info(self, rq_info_path=GET_DATA_INFO_PATH):
        # read the request "getData" info
        rq_info = str() # request info
        with open(rq_info_path, "r", encoding = "utf-8") as file:
            rq_info = file.read()
            rq_info = json.loads(rq_info)

        # extract data from the rq_info
        URL = rq_info["url"]
        headers = dict()
        for data in rq_info["headers"]:
            headers[data["name"]] = data["value"]

        postData = dict()
        for i, post in enumerate(rq_info["postData"]["params"]):
            postData[post["name"]] = post["value"]
            
        self.rq_info = Request_info(URL=URL, headers=headers, postData=postData)
        
    def _adjust_postData(self, request_param, economy:str, product:str, partner:str):
        # The transmission from str to code occurs in this function
        postData = self.rq_info.postData
        reportView = unquote(postData["sWD_ReportView"])
        reportView = '<?xml version="1.0" encoding="UTF-8"?>' + reportView
        reportView = reportView.replace('+', ' ') 
        reportView = ET.fromstring(reportView)
        
        # request specific data
        postData["sWD_ReportId"] = request_param.sWD_ReportId
        postData["sWD_TableId"] = request_param.sWD_TableId
        
        # adjust economy
        for target in reportView.findall(r"./RowDims/Dim[@name='ECONOMY']"):
            target.find(".//String").attrib["value"] = economy # only one each request
        
        # adjust product
        for target in reportView.findall(r"./OtherDims/Dim[@name='PRODUCT']"):
            target.find(".//ActiveItem").attrib["pos"] = product # only one each request
        
        # adjust partner
        for target in reportView.findall(r"./OtherDims/Dim[@name='PARTNER']"):
            target.find(".//String").attrib["value"] = partner # only one each request
        
        reportView = ET.tostring(reportView).decode("utf-8")
        postData["sWD_ReportView"] = reportView
        postDataStr = ""
        count = 0
        for key, value in postData.items():
            if(count != 0):
                postDataStr = postDataStr + '&'
            postDataStr = postDataStr + key + '=' + value
            count += 1
        return postDataStr
        
    def _specific_request(self, request_param, economy:str, product:str, partner:str):
        postData = self._adjust_postData(request_param, economy, product, partner)
        dataGetter = requests.post(url=self.rq_info.URL, headers=self.rq_info.headers, data=postData)
        target = list()
        result = ET.fromstring(dataGetter.text)
        for value in result.findall(".//C"):
            try:
                target.append(int("".join(value.attrib["f"].split()))) # get rid of space
            except ValueError:
                target.append(value.attrib["f"]) # deal with NA
        return target
        
    def _save_one_sheet(self, ws, 
                        request_param:Request_param, 
                        economy, Product, Partner, Period, 
                        display_progress_rate=True):
        
        # economy, time, partner, product_group
        ws.title = request_param.sheet_name
        ws["A1"] = request_param.sheet_title
        ws["A4"] = "ECONOMY"
        ws["A6"] = "YEAR"
        ws["A7"] = "PRODUCT"
        ws["B7"] = "PARTNER"
        
        line = 8
        
        ws["B4"] = economy.name
        for product in Product:
            cellColumn = ord('C')
            ws["A{}".format(line)] = product.name
            for partner in Partner:
                ws["B{}".format(line)] = partner.name
                target = self._specific_request(request_param, economy.code, product.code, partner.code)
                
                # deal with the fixed year period
                if(line == 8):
                    startYear = 1995
                    ws["A1"] = ws["A1"].value + "{}-{}".format(startYear, startYear+len(target)-1)
                    for i in range(len(target)):
                        ws.cell(row=6, column=3+i, value=str(startYear+i))
                
                for i in range(len(target)):
                    ws["{}{}".format(chr(cellColumn+i), line)] = target[i]
                    time.sleep(self.delay_max * random())
                line += 1
            if(display_progress_rate):
                print("Finish saving table: category-\"{}\", economy-\"{}\", product-\"{}\"".
                      format(request_param.sheet_name, economy.name.strip(), product.name.strip()))
    
    def request_and_save(self, whatData:str, Economy:list, Product:list, Partner:list,
                        output_path:str, Period = None, display_progress_rate=True):
        '''
        @param whatData: What data do you want to get? We support "import",
                        "export" , and "both"
        @param Economy: What economy do you want to know?
        @param Product: What product do you want to know?
        @param Partner: What partner do you want to know?
        @param output_path: After crawling the ordered data, where should we store them
        @param display_progress_rate:
            If this parameter is set to True, then you can track the progress of
            the crawler when it start crawling.

        @param Period: not support yet.

        Note: when the number of Economy is more than 1, the crawler will store the
            data into saperate files. Their names whould become
            output_path_<Economy_name>.xlsx.  
        '''
        if(whatData not in SUPPORT_DATA):
            raise ValueError("[!] The data you request is not supported.  ")

        # This block may move to _save_one_ws to get more flexibility in the fulture, 
        # though put it here can increase some efficiency
        Economy = tuple([self.economyCode.search(i) for i in Economy])
        Product = tuple([self.productCode.search(i) for i in Product])
        Partner = tuple([self.partnerCode.search(i) for i in Partner])
        # ---
        
        category = CATEGORY[whatData.lower()]
        for economy in Economy:
            wb = Workbook()
            ws = wb.active
            for i, param in enumerate(category):
                if(i != 0):
                    ws = wb.create_sheet()
                self._save_one_sheet(ws, param, economy, Product, Partner, Period, 
                                     display_progress_rate=display_progress_rate)
            
            adjust_output_path = output_path
            folder, file = splitPath(output_path)

            if(not isdir(folder)):
                mkdir(folder)
            if(len(Economy) != 1):
                fileName, ext = file.split('.')
                adjust_fileName = "{}_{}".format(fileName, economy.name.strip().replace(' ', '_'))
                adjust_output_path = joinPath(folder, "{}.{}".format(adjust_fileName, ext))
            
            wb.save(adjust_output_path)
            if(display_progress_rate):
                print("Finish saving the data of economy:{}. ".format(economy.name.strip()))
    
    def update_economy(self):
        self.economyCode.update()
    
    def update_product(self):
        self.productCode.update()
    
    def update_partner(self):
        self.partnerCode.update()
    
    def update_all_code(self):
        self.update_economy()
        time.sleep(random() * self.delay_max)
        
        self.update_product()
        time.sleep(random() * self.delay_max)
        
        self.update_partner()
        time.sleep(random() * self.delay_max)
        
    def save_all_code(self):
        self.economyCode.save()
        self.productCode.save()
        self.partnerCode.save()